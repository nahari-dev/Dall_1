"""Extract questions from dall_academy_sdle_dataset.xlsx into data/seed_questions.json."""

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Run: pip install openpyxl")

XLSX_PATH = Path("/Users/alinaahari/Downloads/dall_academy_sdle_dataset.xlsx")
OUT_PATH = Path(__file__).parent.parent / "data" / "seed_questions.json"

SECTION_MAP = {
    "Endodontics": "Endodontics",
    "Fixed Prosthodontics": "Fixed Prosthodontics",
    "Implantology": "Implantology",
    "Oral Medicine / Oral Surgery / Medically Compromised Patients": "Oral Medicine & Surgery",
    "Orthodontics": "Orthodontics",
    "Pediatric Dentistry": "Pediatric Dentistry",
    "Periodontics": "Periodontics",
    "Professionalism / Infection Control / Patient Safety": "Professionalism & Safety",
    "Removable Prosthodontics": "Removable Prosthodontics",
    "Restorative": "Restorative Dentistry",
}

wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb["Deduplicated"]

questions = []
for r in range(2, ws.max_row + 1):
    row = [ws.cell(r, c).value for c in range(1, 13)]
    dataset_id, section, question_text = row[0], row[1], row[2]
    opt_a, opt_b, opt_c, opt_d, opt_e, opt_f = row[3], row[4], row[5], row[6], row[7], row[8]
    source_file, page = row[9], row[10]

    if not question_text:
        continue

    options = {}
    for key, val in [("A", opt_a), ("B", opt_b), ("C", opt_c), ("D", opt_d), ("E", opt_e), ("F", opt_f)]:
        if val is not None:
            options[key] = str(val).strip()

    num = re.sub(r"\D", "", str(dataset_id))
    qid = f"q{int(num):04d}"

    questions.append({
        "id": qid,
        "topic": SECTION_MAP.get(section, section),
        "subtopic": "",
        "difficulty": "medium",
        "text": str(question_text).strip(),
        "options": options,
        "correct_answer": None,
        "explanation": "",
        "citations": [],
        "sdle_year": "2025",
        "source_file": str(source_file) if source_file else "",
        "source_page": int(page) if page else None,
    })

OUT_PATH.write_text(json.dumps(questions, indent=2, ensure_ascii=False), encoding="utf-8")
print(f"Wrote {len(questions)} questions to {OUT_PATH}")

# Print topic breakdown
from collections import Counter
counts = Counter(q["topic"] for q in questions)
for topic, count in sorted(counts.items(), key=lambda x: -x[1]):
    print(f"  {topic}: {count}")
